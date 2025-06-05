# -*- coding: utf-8 -*-
import os
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib import colors
from reportlab.lib.units import inch
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
import io
import matplotlib.pyplot as plt # Importar para salvar a figura

def exportar_para_pdf(report_data, chart_figure, output_path):
    """Gera um relatório em PDF com os dados e o gráfico."""
    if not report_data:
        raise ValueError("Dados do relatório não disponíveis para exportação.")

    doc = SimpleDocTemplate(output_path, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Título
    title = Paragraph("Relatório de Vendas Diário", styles["h1"])
    title.alignment = TA_CENTER
    story.append(title)
    story.append(Spacer(1, 0.2*inch))

    # Data de Geração
    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    date_text = Paragraph(f"Gerado em: {now}", styles["Normal"])
    date_text.alignment = TA_CENTER
    story.append(date_text)
    story.append(Spacer(1, 0.3*inch))

    # Dados Principais
    story.append(Paragraph("<b>Resumo das Vendas:</b>", styles["h2"]))
    story.append(Spacer(1, 0.1*inch))

    data_summary = [
        [Paragraph("Faturamento Total:", styles["Normal"]), Paragraph(f"R$ {report_data["faturamento"]:.2f}", styles["Normal"])],
        [Paragraph("Produto Mais Vendido:", styles["Normal"]), Paragraph(str(report_data["mais_vendido"]), styles["Normal"])],
        [Paragraph("Total de Itens Vendidos:", styles["Normal"]), Paragraph(str(report_data["total_itens"]), styles["Normal"])],
    ]
    summary_table = Table(data_summary, colWidths=[2*inch, 4*inch])
    summary_table.setStyle(TableStyle([
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        # (")GRID", (0,0), (-1,-1), 1, colors.grey), # Descomente para ver as bordas
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))

    # Gráfico
    if chart_figure:
        story.append(Paragraph("<b>Gráfico - Top Produtos Mais Vendidos (Quantidade):</b>", styles["h2"]))
        story.append(Spacer(1, 0.1*inch))
        
        # Salvar figura em memória para usar no PDF
        img_buffer = io.BytesIO()
        chart_figure.savefig(img_buffer, format=\'png\
', dpi=300, bbox_inches=\'tight\')
        img_buffer.seek(0)
        
        # Ajustar tamanho da imagem no PDF
        img_width = 6 * inch
        aspect = chart_figure.get_figheight() / chart_figure.get_figwidth()
        img_height = img_width * aspect

        img = Image(img_buffer, width=img_width, height=img_height)
        img.hAlign = \'CENTER\'
        story.append(img)
        story.append(Spacer(1, 0.2*inch))

    # Tabela de Vendas por Produto (se disponível)
    vendas_por_produto = report_data.get("vendas_por_produto")
    if vendas_por_produto is not None and not vendas_por_produto.empty:
        story.append(Paragraph("<b>Vendas por Produto (Quantidade):</b>", styles["h2"]))
        story.append(Spacer(1, 0.1*inch))

        # Preparar dados para a tabela
        table_data = [["Produto", "Quantidade Vendida"]]
        for produto, quantidade in vendas_por_produto.items():
            table_data.append([Paragraph(str(produto), styles["Normal"]), Paragraph(str(quantidade), styles["Normal"])])

        product_table = Table(table_data, colWidths=[4*inch, 2*inch])
        product_table.setStyle(TableStyle([
            (\'BACKGROUND\', (0,0), (-1,0), colors.grey),
            (\'TEXTCOLOR\', (0,0), (-1,0), colors.whitesmoke),
            (\'ALIGN\', (0,0), (-1,-1), \'CENTER\'),
            (\'FONTNAME\', (0,0), (-1,0), \'Helvetica-Bold\'),
            (\'BOTTOMPADDING\', (0,0), (-1,0), 12),
            (\'BACKGROUND\', (0,1), (-1,-1), colors.beige),
            (\'GRID\', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(product_table)

    # Construir o PDF
    try:
        doc.build(story)
    except Exception as e:
        raise Exception(f"Erro ao construir o PDF: {e}")

def exportar_para_excel(report_data, dataframe, chart_figure, output_path):
    """Gera um relatório em Excel com os dados e o gráfico."""
    if not report_data or dataframe is None:
        raise ValueError("Dados do relatório ou DataFrame não disponíveis para exportação.")

    try:
        workbook = openpyxl.Workbook()
        sheet_resumo = workbook.active
        sheet_resumo.title = "Resumo"

        # Escrever Resumo
        sheet_resumo["A1"] = "Relatório de Vendas Diário"
        sheet_resumo["A1"].font = openpyxl.styles.Font(bold=True, size=14)
        sheet_resumo["A2"] = f"Gerado em: {datetime.now().strftime(\"%d/%m/%Y %H:%M:%S\")}"

        sheet_resumo["A4"] = "Faturamento Total:"
        sheet_resumo["B4"] = report_data["faturamento"]
        sheet_resumo["B4"].number_format = \'R$ #,##0.00\'
        sheet_resumo["A5"] = "Produto Mais Vendido:"
        sheet_resumo["B5"] = report_data["mais_vendido"]
        sheet_resumo["A6"] = "Total de Itens Vendidos:"
        sheet_resumo["B6"] = report_data["total_itens"]

        # Adicionar dados brutos em outra aba
        sheet_dados = workbook.create_sheet(title="Dados Completos")
        # Escrever cabeçalhos
        for col_num, column_title in enumerate(dataframe.columns, 1):
            cell = sheet_dados.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = openpyxl.styles.Font(bold=True)
        # Escrever linhas de dados
        for row_num, row_data in enumerate(dataframe.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                sheet_dados.cell(row=row_num, column=col_num).value = cell_value
                # Formatar data se for datetime
                if isinstance(cell_value, (datetime, pd.Timestamp)):
                     sheet_dados.cell(row=row_num, column=col_num).number_format = \'DD/MM/YYYY\'
                # Formatar valor total
                if dataframe.columns[col_num-1] == \'valor_total\':
                     sheet_dados.cell(row=row_num, column=col_num).number_format = \'R$ #,##0.00\'


        # Adicionar gráfico (como imagem)
        if chart_figure:
            img_buffer = io.BytesIO()
            chart_figure.savefig(img_buffer, format=\'png\
', dpi=150, bbox_inches=\'tight\')
            img_buffer.seek(0)
            img = OpenpyxlImage(img_buffer)
            img.anchor = \'A8\'
            sheet_resumo.add_image(img)

        # Salvar o arquivo
        workbook.save(output_path)

    except Exception as e:
        raise Exception(f"Erro ao gerar o arquivo Excel: {e}")
